import os
from typing import final

import colorama
from colorama import init, Fore, Back, Style
import warnings
from com_utils import find_available_ports, send_at_command
colorama.init()
# Находим все доступные COM порты
available_ports = find_available_ports()

if not available_ports:
    print(Fore.LIGHTWHITE_EX+"Не удалось найти модем.")
    print('Функции отправки и принятия СМС не будут работать.', Fore.LIGHTWHITE_EX)

else:
    num_ports = len(available_ports)
    # Проверяем настройки отладки из файла settings.txt
    settings_file = "Files/settings.txt"
    debug_mode = False
    if os.path.exists(settings_file):
        with open(settings_file, 'r') as file:
            for line in file:
                if line.strip() == 'debug = 1':
                    debug_mode = True
                    break

    # Проходим по каждому доступному порту
    for port in available_ports:
        #if debug_mode:
            #print(f"Отправка AT команды на порт {port}...        debug")
        response = send_at_command(port)
        if response:
            #if debug_mode:
                #print(f"Ответ от порта {port}: {response}")
            # Сохраняем первый найденный порт и завершаем выполнение
            modem_port = port
            break
        if not available_ports:
            modem_port = 'COM'
from datetime import timedelta
warnings.simplefilter(action='ignore', category=FutureWarning)

def load_contacts(filename):
    print("what")
    try:
        df = pd.read_excel(filename)
        df = df.drop_duplicates()
    except Exception as e:
        print(f"Ошибка при чтении файла контактов: {e}")
        return {}

    required_columns = [df.columns[0], df.columns[1]]  # Номер телефона, Имя
    for column in required_columns:
        if column not in df.columns:
            print(f"Отсутствует ожидаемый столбец: '{column}'")
            return {}

    contacts = {}
    for index, row in df.iterrows():
        phone_number = str(row[df.columns[0]]).replace("а",'%').replace(' ', '').replace('-', '')
        print(f"Phone nu = {phone_number}")
        contacts[phone_number] = phone_number #row[df.columns[1]]
    return contacts
def load_sms_log(filename):
    try:
        df = pd.read_excel(filename)
        df = df.drop_duplicates()
        df['Сообщение'] = df['Сообщение'].astype(str)  # Преобразуем столбец 'Сообщение' в строковый тип
        print("Загруженные столбцы:", df.columns)  # Вывод загруженных столбцов для отладки
    except Exception as e:
        print(f"Ошибка при чтении файла SMS логов: {e}")
        return pd.DataFrame()
    return df


def delete_contact(nums):
    ii=0
    try:
            # Загружаем файл
            wb = load_workbook("Files/contacts.xlsx")
            ws = wb.active
            # Находим все строки для удаления (в обратном порядке)
            rows_to_delete = []
            for row in range(ws.max_row, 1, -1):  # начинаем с конца, пропускаем заголовок
                if f"+7{ws.cell(row=row, column=1).value}" in nums:
                    ii+=1
                    rows_to_delete.append(row)

            # Удаляем найденные строки
            for row_idx in rows_to_delete:
                ws.delete_rows(row_idx, 1)

            # Сохраняем изменения
            wb.save("Files/contacts.xlsx")
            return True, f"{ii} Контактов успешно удалено"

    except Exception as e:
        return False, f"Ошибка при удалении контакта: {str(e)}"


def get_current_datetime():
    now = datetime.now()
    return now.strftime('%d/%m/%Y'), now.strftime('%H:%M:%S')
def analyze_sms_log(contacts_file, sms_log_file, analysis_file):
    contacts = load_contacts(contacts_file)
    sms_log = load_sms_log(sms_log_file)

    if sms_log.empty or not contacts:
        print("Не удалось загрузить данные для анализа.")
        return

    # Получение даты и времени
    today_date, current_time = get_current_datetime()
    yesterday_date = (datetime.now() - timedelta(days=1)).strftime('%d/%m/%Y')

    # Используем индексацию по номерам столбцов
    phone_column_index = 0    # Индекс столбца с номерами телефонов
    date_column_index = 3     # Индекс столбца с датой получения SMS

    # Проверка типов данных
    sms_log.iloc[:, phone_column_index] = sms_log.iloc[:, phone_column_index].astype(str)
    sms_log.iloc[:, date_column_index] = pd.to_datetime(sms_log.iloc[:, date_column_index], format='%d/%m/%Y', errors='coerce').dt.strftime('%d/%m/%Y')

    # Получаем номера телефонов, которые отправляли SMS за последние сутки
    recent_sms = sms_log[sms_log.iloc[:, date_column_index].isin([today_date, yesterday_date])]
    recent_sms_numbers = recent_sms.iloc[:, phone_column_index].str.replace(' ', '').str.replace('-', '').str.replace('+', '').unique()

    # Определяем контакты, которые не прислали SMS за последние сутки
    missing_contacts = {number: name for number, name in contacts.items() if number not in recent_sms_numbers}

    # Определяем номер анализа
    try:
        with open(analysis_file, 'r', encoding='utf-8') as f:
            lines = f.readlines()
            last_analysis_number = None
            for line in reversed(lines):
                if line.startswith('Анализ номер '):
                    last_analysis_number = int(line.strip().split('Анализ номер ')[-1].split()[0])
                    break
            if last_analysis_number is None:
                new_analysis_number = 1
            else:
                new_analysis_number = last_analysis_number + 1
    except FileNotFoundError:
        new_analysis_number = 1

    # Запись результатов в файл
    analysis_content = f"Анализ номер {new_analysis_number}\n"
    analysis_content += f"Дата: {today_date}, время: {current_time}.\n"
    analysis_content += "Контакты, не приславшие сообщение за последние сутки:\n\n"
    for number, name in missing_contacts.items():
        if len(number) == 11 and not number.startswith('8'):
            analysis_content += f"+{number} -- {name}\n"
        else:
            analysis_content += f"{number} -- {name}\n"

    # Анализ сообщений
    settings = read_settings("Files/settings.txt")
    charge_warning = int(settings.get('charge_warning', 0))
    wb = load_workbook("Files/sms_log.xlsx", data_only=True)
    ws = wb.active

    # Очистка 7-го столбца
    for row in ws.iter_rows(min_row=2, values_only=False):
        row[5].value = ""

    # Добавляем столбец "Отклонения", если его нет
    if ws.max_column < 6:
        ws['G1'] = 'Отклонения'

    for row in ws.iter_rows(min_row=2, values_only=False):
        message = row[2].value
        if isinstance(message, str):  # Проверяем, является ли сообщение строкой
            deviations = row[5].value if row[5].value else ""
            battery_warning = False
            gps_warning = False
            for line in message.splitlines():
                if "Спутн: 0" in line:
                    gps_warning = True
                if "Бат:" in line:
                    battery_level = int(line.split("(")[1].split("%")[0])
                    if battery_level < charge_warning:
                        battery_warning = True
            if battery_warning:
                deviations += "Бат! "
            if gps_warning:
                deviations += "GPS! "
            row[5].value = deviations

            # Устанавливаем цвет фона для 7-го столбца
            if "Бат! GPS! " in deviations:
                row[5].fill = PatternFill(start_color='FFFF950E', end_color='FFFF950E', fill_type='solid')
            elif "GPS! " in deviations:
                row[5].fill = PatternFill(start_color='FFF0F076', end_color='FFF0F076', fill_type='solid')
            elif "Бат! " in deviations:
                row[5].fill = PatternFill(start_color='FFAFEEEE', end_color='FFAFEEEE', fill_type='solid')

    # Установка белой заливки для пустых ячеек в 7-м столбце
    for row in ws.iter_rows(min_row=2, values_only=False):
        if row[5].value is None or row[5].value == "":
            row[5].fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

    wb.save(sms_log_file)

    with open(analysis_file, 'a', encoding='utf-8') as f:
        f.write("\n\n")
        f.write(analysis_content)

    print("Анализ:")
    print(analysis_content)
    print(f"Анализ номер {new_analysis_number} успешно добавлен в файл {analysis_file}.")
def analysis():
    contacts_file = "Files/contacts.xlsx"
    sms_log_file = "Files/sms_log.xlsx"
    analysis_file = "Files/Analysis.txt"
    analyze_sms_log(contacts_file, sms_log_file, analysis_file)
def clear_console():
    # Определяем операционную систему
    current_os = platform.system()

    # Очищаем консоль в зависимости от ОС
    if current_os == 'Windows':
        os.system('cls')
    elif current_os in ['Linux', 'Darwin']:  # Darwin - это macOS
        os.system('clear')
    else:
        print("Операционная система не поддерживается для очистки консоли.")
def add_contacts(file_path, new_contacts):
    # Создаем каталог, если он не существует
    directory = os.path.dirname(file_path)
    if not os.path.exists(directory):
        os.makedirs(directory)

    if os.path.exists(file_path):
        # Загружаем существующий файл
        wb = load_workbook(file_path)
        ws = wb.active
    else:
        # Создаем новый файл и заполняем заголовки
        wb = Workbook()
        ws = wb.active
        ws.title = "Contacts"
        ws.append(["Phone Number", "Contact Name"])

    # Добавляем новые контакты
    for contact in new_contacts:
        ws.append(contact)

    wb.save(file_path)
    print(f"Контакты успешно добавлены в {file_path}")
def send_smst():
    contacts_file = "Files/contacts.xlsx"
    sms_message = input("Введите сообщение (английскими буквами!): ")
    search_terms = input("Введите аргументы для поиска: ")
    search_terms = search_terms.split()
    include_terms = [term for term in search_terms if not term.startswith('-')]
    exclude_terms = [term[1:] for term in search_terms if term.startswith('-')]

    wb = load_workbook(contacts_file)
    ws = wb.active

    contacts_to_send = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        phone_number, contact_name = row
        if not search_terms:
            contacts_to_send.append((phone_number, contact_name))
        elif (any(term in phone_number or term in contact_name for term in include_terms) and
              not any(term in phone_number or term in contact_name for term in exclude_terms)):
            contacts_to_send.append((phone_number, contact_name))

    if not contacts_to_send:
        print("Нет контактов, соответствующих критериям поиска.")
        return

    print("Найдены следующие контакты:")
    for i, contact in enumerate(contacts_to_send):
        print(f"{i+1}. {contact[0]} -- {contact[1]}")

    while True:
        confirm = input("Нажмите Enter для подтверждения: ")
        if confirm.lower() == "":
            for contact in contacts_to_send:
                send_sms(modem_port, contact[0], sms_message, 'text', debug=False)
            break
        elif confirm.lower() == "e":
            print("Текущие аргументы:")
            print(f"Сообщение: {sms_message}")
            print(f"Поиск: {' '.join(search_terms)}")
            new_sms_message = input("Введите новое сообщение (английскими буквами!): ")
            new_search_terms = input("Введите новые номера телефонов или имена контактов для отправки сообщения (через пробел, оставьте пустым для всех контактов): ")
            sms_message = new_sms_message if new_sms_message else sms_message
            search_terms = new_search_terms.split() if new_search_terms else search_terms
            include_terms = [term for term in search_terms if not term.startswith('-')]
            exclude_terms = [term[1:] for term in search_terms if term.startswith('-')]
            contacts_to_send = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                phone_number, contact_name = row
                if not search_terms:
                    contacts_to_send.append((phone_number, contact_name))
                elif (any(term in phone_number or term in contact_name for term in include_terms) and
                      not any(term in phone_number or term in contact_name for term in exclude_terms)):
                    contacts_to_send.append((phone_number, contact_name))
            print("Найдены следующие контакты:")
            for i, contact in enumerate(contacts_to_send):
                print(f"{i+1}. {contact[0]} -- {contact[1]}")
        elif confirm.lower() == "n":
            break
        else:
            print("Недопустимый выбор. Пожалуйста, выберите действие из меню.")
def delete_contacts(file_path, search_terms):
    wb = load_workbook(file_path)
    ws = wb.active

    include_terms = [term for term in search_terms if not term.startswith('-')]
    exclude_terms = [term[1:] for term in search_terms if term.startswith('-')]

    contacts_to_delete = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        phone_number, contact_name = row
        if not search_terms:
            contacts_to_delete.append((phone_number, contact_name))
        elif (any(term in phone_number or term in contact_name for term in include_terms) and
              not any(term in phone_number or term in contact_name for term in exclude_terms)):
            contacts_to_delete.append((phone_number, contact_name))

    if not contacts_to_delete:
        print("Нет контактов, соответствующих критериям поиска.")
        return

    print("Найдены следующие контакты:")
    for i, contact in enumerate(contacts_to_delete):
        print(f"{i+1}. {contact[0]} -- {contact[1]}")

    while True:
        confirm = input("Нажмите Enter для подтверждения: ")
        if confirm.lower() == "":
            for contact in contacts_to_delete:
                for row in ws.iter_rows(min_row=2, values_only=False):
                    if row[0].value == contact[0] and row[1].value == contact[1]:
                        ws.delete_rows(row[0].row)
            wb.save(file_path)
            print(f"Контакты успешно удалены из {file_path}")
            break
        elif confirm.lower() == "e":
            print("Текущие аргументы:")
            print(f"Поиск: {' '.join(search_terms)}")
            new_search_terms = input("Введите новые номера телефонов или имена контактов для удаления (через пробел, оставьте пустым для всех контактов): ")
            search_terms = new_search_terms.split() if new_search_terms else search_terms
            include_terms = [term for term in search_terms if not term.startswith('-')]
            exclude_terms = [term[1:] for term in search_terms if term.startswith('-')]
            contacts_to_delete = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                phone_number, contact_name = row
                if not search_terms:
                    contacts_to_delete.append((phone_number, contact_name))
                elif (any(term in phone_number or term in contact_name for term in include_terms) and
                      not any(term in phone_number or term in contact_name for term in exclude_terms)):
                    contacts_to_delete.append((phone_number, contact_name))
            print("Найдены следующие контакты:")
            for i, contact in enumerate(contacts_to_delete):
                print(f"{i+1}. {contact[0]} -- {contact[1]}")
        elif confirm.lower() == "n":
            break
        else:
            print("Недопустимый выбор. Пожалуйста, выберите действие из меню.")
def search_contacts(file_path, search_terms):
    wb = load_workbook(file_path)
    ws = wb.active

    include_terms = [term for term in search_terms if not term.startswith('-')]
    exclude_terms = [term[1:] for term in search_terms if term.startswith('-')]

    if search_terms == "":
        search_terms = ["9","8","7","6","5","4","3","2","1","0"]
    else:
        search_terms = search_terms.split()


    final_strings = []

    for search_term in search_terms:
        if not search_term.startswith("-"):
            for row in ws.iter_rows(min_row=2, values_only=True):
                phone_number, contact_name = row
                if phone_number:
                    string = f"+7{phone_number}::{contact_name}".replace("    ","")

                    if search_term in string and string not in final_strings:
                        final_strings.append(string)
        else:
            argument = search_term.replace("-","")
            for final_string  in final_strings:
                if argument in final_string:
                    final_strings.remove(final_string)


    contacts_found = []

    for final_string in final_strings:
        num1, name1 = final_string.split("::")
        contacts_found.append({"num":num1,"name":name1})

    if not contacts_found:
        print("Нет контактов, соответствующих критериям поиска.")
        return ["Нет контактов, соответствующих критериям поиска."], []

    print(f"{contacts_found=}")
    final = []
    print("Найдены следующие контакты:")

    just_info = []

    for i, contact in enumerate(contacts_found):
        just_info.append({"number":contact["num"], "name": contact["name"]})
        string = f"{i+1}. {contact["num"]} -- {contact["name"]}"
        final.append(string)
        print(string)
    print(f"{final=}")
    return final, just_info

def edit_contacts():
    file_path = "Files/contacts.xlsx"

    while True:
        print("\nМеню:")
        print()
        print("1. Добавить контакт;")
        print("2. Удалить контакт;")
        print("3. Поиск контактов;")
        print("4. Выход;")
        print()
        choice = str(input("Выберите действие: "))

        if choice == str("1"):
            new_contacts = []
            while True:
                phone_number = input("Введите номер телефона (Оставить пустым для завершения): ")
                if not phone_number:
                    break
                contact_name = input("Введите имя контакта: ")
                new_contacts.append([phone_number, contact_name])
            if new_contacts:
                add_contacts(file_path, new_contacts)
            else:
                print("Нет контактов для добавления.")
        elif choice == str("2"):
            search_terms = input("Аргументы для поиска: ")
            search_terms = search_terms.split()
            delete_contacts(file_path, search_terms)
        elif choice == str("3"):
            search_terms = input("Аргументы для поиска: ")
            search_terms = search_terms.split()
            search_contacts(file_path, search_terms)
        elif choice == "4":
            break
        else:
            print("Недопустимый выбор. Пожалуйста, выберите действие из меню.")
            break
import subprocess
import platform
def open_files_folder():
    """
    Открывает папку 'Files' в текущем каталоге в проводнике.
    """
    try:
        # Определяем путь к папке 'Files'
        current_directory = os.getcwd()
        folder_path = os.path.join(current_directory, 'Files')

        # Проверяем, существует ли папка
        if not os.path.isdir(folder_path):
            print(f"Папка не существует: {folder_path}")
            return

        # Открываем папку в проводнике
        if os.name == 'nt':  # Windows
            subprocess.run(['explorer', folder_path], check=True)
        elif os.name == 'posix':  # macOS or Linux
            if sys.platform == 'darwin':  # macOS
                subprocess.run(['open', folder_path], check=True)
            else:  # Linux
                subprocess.run(['xdg-open', folder_path], check=True)
        else:
            print(f"Операционная система не поддерживается: {os.name}")

    except subprocess.CalledProcessError as e:
        print('', end = '')
    except Exception as e:
        print('', end = '')
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
def send_at_command0(ser, command, response_timeout=1):
    ser.write((command + '\r\n').encode())
    time.sleep(response_timeout)
    response = ser.read_all().decode()
    return response
def send_at_command(port, command):
    modem = serial.Serial(port, 9600, timeout=5)
    modem.write((command + '\r\n').encode())
    time.sleep(1)
    response = modem.read_all().decode()
    modem.close()
    return response
def delete_sms_by_index(port, index):
    try:
        modem = serial.Serial(port, 9600, timeout=5)
        time.sleep(1)
        modem.write(f'AT+CMGD={index}\r\n'.encode())  # Удаляем сообщение по индексу
        time.sleep(1)
        modem.close()
    except serial.SerialException as e:
        print(f"Ошибка открытия порта {port}: {e}")
    except Exception as e:
        print(f"Ошибка при удалении смс по индексу {index}: {e}")
def format_date(date_str):
    try:
        # Предполагается, что дата в формате YY/MM/DD
        date_obj = datetime.strptime(date_str, '%y/%m/%d')
        return date_obj.strftime('%d/%m/%Y')
    except ValueError:
        return date_str  # В случае ошибки возвращаем оригинальную строку
# Функция для парсинга ответа AT+CMGL и извлечения SMS сообщений
def parse_sms_response(response):
    messages = []
    lines = response.splitlines()
    i = 0
    '''
    пример сообщения: 
    +CMGL: 10,"REC READ","+79875324724",,"24/11/15,14:35:51+12"
    Hello!
    Или:

    '''
    while i < len(lines):
        if "+CMGL: " in lines[i]:
            parts = lines[i].split(",")
            index = parts[0].split(": ")[1].strip()
            sender_number = parts[2].strip('"')
            date_and_time = lines[i].split(",,")[1].replace('"','').split(',')
            print(f"{date_and_time=}")
            date_dates = date_and_time[0].split("/")
            date_date = f"{date_dates[2]}.{date_dates[1]}.{date_dates[0]}"
            #print(f"ДАТА = {date_date}")

            date_time = date_and_time[1].split("+")[0].split("-")[0]
            #print(f"ВРЕМЯ = {date_time}")


            message_lines = []

            # Проверяем, есть ли следующая строка
            if i + 1 < len(lines):
                # Если следующая строка не начинается с "+CMGL: ", то добавляем ее к сообщению
                j = i + 1
                while j < len(lines) and "+CMGL: " not in lines[j]:
                    if "OK" in lines[j]:
                        break
                    message_lines.append(lines[j].strip())
                    j += 1
                i = j - 1  # Установим индекс на последнюю строку сообщения

            # Декодируем строки сообщения
            decoded_lines = []
            for line in message_lines:
                try:
                    # Преобразуем сообщение в формат UCS2
                    decoded_line = bytes.fromhex(line).decode('utf-16be')
                    decoded_lines.append(decoded_line)
                except (ValueError, UnicodeDecodeError):
                    # Если декодирование не удалось, оставляем все строки сообщения в первоначальном виде
                    decoded_lines = message_lines
                    break

            message = '\n'.join(decoded_lines)



            # Преобразуем дату в формат DD/MM/YYYY
            #formatted_date = format_date(date.strip())

            messages.append({
                "index": index,
                "sender_number": sender_number,
                "date": date_date,
                "time": date_time,
                "message": message.strip()
            })
        i += 1
    return messages
# Функция для объединения длинных сообщений
def combine_long_messages(messages):
    combined_messages = []
    for message in messages:
        combined_messages.append(message)
    return combined_messages

def num_to_name(num):
    wb = load_workbook("Files/contacts.xlsx")
    ws = wb.active
    print(f"Искомый номер: {num}")
    for row in ws.iter_rows(min_row=2, values_only=True):
        phone_number, contact_name = row
        if phone_number:
            phone_number = f"+7{phone_number}"
            print(f"Номер контакта: {phone_number}")

            if phone_number == num:
                return contact_name
    return num

# Изменение функции read_sms_and_save
def read_sms_and_save(port, contacts_file, output_file):
        with serial.Serial(port, 9600, timeout=1) as ser:
            print("Проверяем...")
            response = send_at_command0(ser, 'AT+CMGL="ALL"')

            # Обработка ответа и запись в Excel
            sms_messages = parse_sms_response(response)
            combined_messages = combine_long_messages(sms_messages)

            # Проверяем, существует ли файл с контактами
            if not os.path.exists(contacts_file):
                print(f"Файл {contacts_file} не найден.")
                return

            contacts = load_contacts(contacts_file)

            # Вывод содержимого SMS
            if combined_messages:
                print()
                #print("Найдены SMS сообщения:", end = '')
                log = ""
                for sms in combined_messages:
                    #print('')
                    log += f"{num_to_name(sms['sender_number'])}: {sms['message']}  {sms['time']}\n"
                append_to_excel(combined_messages, contacts, output_file)
                #print("Добавлено, удаляем")
                # Удаление SMS по индексу
                for sms in combined_messages:
                    print(f"удаляем {sms}")
                    send_at_command0(ser, f"AT+CMGD={sms['index']}")
                return log
            else:
                cy = 1
                if cy == 15:
                    cy = 1
                return ""

# Функция для загрузки контактов из файла Excel
def load_contacts(filename):
    try:
        df = pd.read_excel(filename)
    except Exception as e:
        print(f"Ошибка при чтении файла контактов: {e}")
        return {}

    required_columns = ['Номер телефона', 'Имя маячка']
    for column in required_columns:
        if column not in df.columns:
            print(f"Отсутствует ожидаемый столбец: '{column}'")
            return {}

    contacts = {}
    for index, row in df.iterrows():
        print(row['Номер телефона'])
        # Приведение номеров телефонов к строковому формату без лишних символов
        phone_number = str(row['Номер телефона']).replace(' ', '').replace('-', '')
        contacts[phone_number] = row['Имя маячка']
    return contacts
from openpyxl.styles import Alignment, PatternFill
def append_to_excel(sms_messages, contacts, output_file):
    if not sms_messages:  # Если нет новых сообщений, не записываем в таблицу
        return
    try:
        wb = load_workbook("Files/sms_log.xlsx", data_only=True)
        ws = wb.active
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.append(["Номер отправителя", "Имя контакта", "Сообщение", "Дата получения", "Время получения"])

    settings = read_settings("Files/settings.txt")
    sleep_time = int(settings.get('sleep_time', 0))  # Значение sleep_time из файла настроек

    for sms in sms_messages:
        sender_number = sms["sender_number"].replace(' ', '').replace('-', '')  # Убираем только пробелы и дефисы, сохраняем +
        contact_name = num_to_name(sms['sender_number'])
        message = sms["message"] if sms["message"] else "Без текста"
        date_received = sms["date"]
        current_date = datetime.now().strftime('%d/%m/%Y')
        current_time = datetime.now().strftime('%H:%M:%S')

        # Ищем существующую строку с таким же номером и временем
        existing_row = None
        for row in ws.iter_rows(min_row=2, values_only=False):
            if (row[0].value == sender_number and
                    row[3].value == date_received and
                    abs((datetime.strptime(current_time, '%H:%M:%S') - datetime.strptime(row[4].value, '%H:%M:%S')).total_seconds()) <= sleep_time + 30):
                existing_row = row
                break

        if existing_row:
            # Если найдена существующая строка, добавляем к ней новое сообщение
            existing_row[2].value += "\n" + message
            # Увеличиваем высоту строки
            lines = existing_row[2].value.count('\n') + 1
            ws.row_dimensions[existing_row[0].row].height = 13.7
        else:
            # Если не найдена существующая строка, добавляем новую
            ws.append([sender_number, contact_name, message, date_received, current_time])
            # Устанавливаем высоту строки
            lines = message.count('\n') + 1
            ws.row_dimensions[ws.max_row].height = 13.7

    wb.save(output_file)
# Функция для обновления имен контактов в sms_log.xlsx
def update_contact_names(output_file, contacts):
    try:
        wb = wb = load_workbook("Files/sms_log.xlsx", data_only=True)
        ws = wb.active
    except FileNotFoundError:
        print(f"Файл {output_file} не найден.")
        return

    for row in ws.iter_rows(min_row=2, values_only=False):
        sender_number = str(row[0].value).replace(' ', '').replace('-', '').replace('+', '')
        recorded_name = row[1].value
        correct_name = contacts.get(sender_number, "Неизвестный")

        if recorded_name != correct_name:
            row[1].value = correct_name

    wb.save(output_file)
# Функция для удаления всех SMS на SIM-карте
def delete_all_sms(port):
    modem = serial.Serial(port, 9600, timeout=5)
    time.sleep(1)
    modem.write(b'AT+CMGD=1,4\r\n')  # Удаляем все сообщения
    time.sleep(1)
    modem.close()
# Основной код
import signal
import sys

# Основной код
def read_sms_to_excel():
    contacts_file = "Files/contacts.xlsx"  # Путь к файлу с контактами
    output_file = "Files/sms_log.xlsx"
    print('Интенсивный поиск смс!')

    # Читаем значение sleep_time из файла settings.txt
    sleep_time = None
    with open("Files/settings.txt", 'r') as file:
        for line in file:
            if line.startswith('sleep_time = '):
                sleep_time = int(line.strip().split(' = ')[1])
                break

    if sleep_time is None:
        print("Не удалось найти настройку sleep_time в файле settings.txt.")
        sleep_time = 0  # Значение по умолчанию

    for i in range(10):
        read_sms_and_save(modem_port, contacts_file, output_file)
        time.sleep(1)
    print("Замедление...")
    while True:
        read_sms_and_save(modem_port, contacts_file, output_file)
        time.sleep(sleep_time)

import serial
import time
from openpyxl import load_workbook
import os
#from com_port_checker import *

import shutil
from datetime import datetime

def clear_logs():
    log_file = "Files/sms_log.xlsx"
    analysis_file = "Files/Analysis.txt"
    archive_dir = "Files/Archive"
    if not os.path.exists(archive_dir):
        os.makedirs(archive_dir)

    current_datetime = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    archived_log_file = f"{archive_dir}/sms_log_{current_datetime}.xlsx"
    archived_analysis_file = f"{archive_dir}/Analysis_{current_datetime}.txt"

    try:
        shutil.copy2(log_file, archived_log_file)
        print(f"Лог успешно архивирован в {archived_log_file}")

        wb = load_workbook("Files/sms_log.xlsx", data_only=True)
        ws = wb.active
        # Удаляем все строки, кроме первой
        ws.delete_rows(2, ws.max_row)

        # Устанавливаем высоту первой строки по умолчанию
        ws.row_dimensions[1].height = 13.7

        # Удаляем все настройки высоты строк, чтобы они стали по умолчанию
        for row_dim in ws.row_dimensions:
            if row_dim!= 1:
                ws.row_dimensions[row_dim].height = None

        wb.save(log_file)
        print("Лог успешно очищен, кроме строки заголовков")

        # Архивирование файла analysis.txt
        shutil.copy2(analysis_file, archived_analysis_file)
        print(f"Файл analysis.txt успешно архивирован в {archived_analysis_file}")

        # Очистка файла analysis.txt
        with open(analysis_file, 'w') as f:
            f.write('')

        print("Файл analysis.txt успешно очищен")

    except Exception as e:
        print(f"Ошибка при очистке логов: {e}")

def find_com_port():
    # Выполняем поиск порта из com_port_checker.py
    # Здесь нужно вставить ваш код для нахождения порта
    return modem_port  # Пример порта, замените на фактический порт, найденный вашим скриптом

def read_settings(settings_file):
    if not os.path.exists(settings_file):
        print(f"Файл {settings_file} не существует.")
        return {}

    settings = {}
    with open(settings_file, 'r') as file:
        for line_num, line in enumerate(file, start=1):
            line = line.strip()
            if not line or '=' not in line:
                continue
            try:
                name, value = line.split('=', 1)
                settings[name.strip()] = value.strip()
            except ValueError as e:
                print(f"Ошибка в строке {line_num}: '{line}', ошибка: {e}")
                continue
    return settings

def text_to_ucs2(text):
    ucs2_text = text.encode('utf-16-be').hex().upper()
    return ucs2_text

def create_pdu_message(phone_number, ucs2_message):
    # Здесь мы создаем PDU-сообщение. Это пример, вам нужно будет адаптировать его под ваши нужды
    service_center_number = "00"  # Используем номер сервисного центра по умолчанию
    tp_mti = "01"  # SMS-SUBMIT
    tp_mr = "00"  # Message reference
    tp_da = "91" + phone_number[1:]  # Номер получателя в международном формате (без '+')
    tp_pid = "00"  # Protocol identifier
    tp_dcs = "08"  # Data coding scheme (UCS2)
    tp_vp = "AA"  # Validity period
    tp_udl = "{:02X}".format(len(ucs2_message) // 2)  # Length of user data
    tp_ud = ucs2_message

    pdu_message = (service_center_number + tp_mti + tp_mr + tp_da + tp_pid + tp_dcs + tp_vp + tp_udl + tp_ud)
    return pdu_message



def send_sms(serial_port, phone_number, message, mode='text', debug=False):
    try:
        # Открываем серийный порт
        ser = serial.Serial(serial_port, 9600, timeout=5)
    except serial.SerialException as e:
        print(f"Ошибка открытия порта {serial_port}: {e}")
        return False

    time.sleep(0.5)  # Ждем немного, чтобы модем успел инициализироваться

    def send_text_mode():
        #ser.write(b'AT+CMGF=1\r')  # Устанавливаем текстовый режим
        time.sleep(0.1)
        ser.write(f'AT+CMGS="{phone_number}"\r'.encode())
        time.sleep(0.5)
        ser.write(message.encode() + b'\x1A')  # Заканчиваем сообщение Ctrl+Z (0x1A)
        time.sleep(0.5)
        return ser.read_all().decode()

    def send_pdu_mode():
        ser.write(b'AT+CMGF=0\r')  # Устанавливаем режим PDU
        time.sleep(1)
        ucs2_message = text_to_ucs2(message)
        pdu_message = create_pdu_message(phone_number, ucs2_message)
        ser.write(f'AT+CMGS={len(pdu_message) // 2}\r'.encode())
        time.sleep(1)
        ser.write(pdu_message.encode() + b'\x1A')  # Заканчиваем сообщение Ctrl+Z (0x1A)
        time.sleep(3)
        return ser.read_all().decode()

    response = send_text_mode()

    if debug:
        print(f"Модем ответил: {response} -debug")

    if 'OK' in response:
        print(f"SMS успешно отправлено на номер {phone_number} ")
        ser.close()
        return True
    else:
        print(f"Ошибка при отправке SMS на номер {phone_number}: {response}")
        ser.close()
        return False

def send_sms_to_contacts(file_path, message):
    if not os.path.exists(file_path):
        print(f"Файл {file_path} не существует.")
        return

    settings = read_settings("Files/settings.txt")
    debug = settings.get('debug') == '1'

    wb = load_workbook(file_path)
    ws = wb.active

    com_port = find_com_port()
    if not com_port:
        print("Не удалось найти COM-порт GSM модема.")
        return

    for row in ws.iter_rows(min_row=2, values_only=True):  # Начинаем со второй строки (первая строка - заголовки)
        phone_number = row[0]
        send_sms(com_port, phone_number, message, 'text', debug)

def restart_modem():
    with serial.Serial(port, 9600, timeout=1) as ser:
        res = send_at_command0(ser, 'AT+CFUN=1,1')
        return True if "OK" in res else False

from com_utils import modem_port

def setup_modem(port):
    with serial.Serial(port, 9600, timeout=1) as ser:
        send_at_command0(ser, 'AT+CMGF=1')
        send_at_command0(ser, 'AT+CPMS="ME","ME","ME"')
        return "OK"

def main():
    if modem_port == 'COM':
        while True:
            print(Fore.LIGHTWHITE_EX+'Hello!')
            print('Функции работы с смс сейчас недоступны.')
            print()
            print('3. Изменить список контактов')
            print('4. Анализировать данные')
            print('5. Открыть папку с файлами')
            print('6. Очистить таблицу')
            print()
            print('Что выполнить?')
            choise = str(input())
            if choise == str(3):
                clear_console()
                edit_contacts()
                input('Enter для выхода')
                clear_console()
            elif choise == str(4):
                clear_console()
                analysis()
                input('Enter для выхода')
                clear_console()
            elif choise == str(5):
                open_files_folder()
                clear_console()
            elif choise == str(6):
                clear_logs()
                input('Enter для выхода ')
                clear_console()
            else:
                clear_console()
                print('Неверный выбор, попробуйте еще раз!')


    else:
        setup_modem(modem_port)
        while True:
            print(Fore.LIGHTWHITE_EX+'Hello!')
            print()
            print('1. Отправка смс')
            print('2. Прием смс и запись в таблицу;')
            print('3. Изменить список контактов;')
            print('4. Анализировать данные;')
            print('5. Открыть папку с файлами;')
            print('6. Очистить таблицу.')
            print()
            print('Что выполнить?')
            choise = str(input())

            if choise == str(1):
                clear_console()
                send_smst()
                input()
                clear_console()
            elif choise == str(2):
                clear_console()
                read_sms_to_excel()
                input()
                clear_console()
            elif choise == str(3):
                clear_console()
                edit_contacts()
                input()
                clear_console()
            elif choise == str(4):
                clear_console()
                analysis()
                input()
                clear_console()
            elif choise == str(5):
                open_files_folder()
                clear_console()
            elif choise == str(6):
                clear_logs()
                input()
                clear_console()
            else:
                clear_console()
                print('Неверный выбор, попробуйте еще раз!')

if __name__ == "__main__":
    main()