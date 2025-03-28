import FreeSimpleGUI as sg
# задаем переменные
modem_port = None
can_modem = False
contacts_data = []
contacts_window = None
model = None
speed = None

def menu_contacts():
    # Куча импорта и глобальных переменных
    global speed
    global model
    global contacts_window
    global can_modem
    global contacts_data
    global modem_port
    import psutil
    import serial.tools.list_ports as list_ports
    import warnings

    contacts_file = "Files/contacts.xlsx"  # Путь к файлу с контактами
    output_file = "Files/sms_log.xlsx"
    settings_file = "Files/settings.txt"
    debug_mode = False

    import os
    def read_settings(settings_file):
        # Функция для чтения настроек

        if not os.path.exists(settings_file):
            print(f"Файл {settings_file}s не существует.")
            os.system(
                """powershell -Command "Add-Type -AssemblyName System.Windows.Forms; [System.Windows.Forms.MessageBox]::Show('Нет директории Files :/', 'Ошибка', [System.Windows.Forms.MessageBoxButtons]::OK, [System.Windows.Forms.MessageBoxIcon]::Error)""")
            exit()

        settings = {}
        with open(settings_file, 'r') as file:
            for line in file:
                # Пропускаем пустые строки и строки без знака '='
                if '=' not in line:
                    continue
                name, value = line.strip().split('=', 1)  # split('=', 1) позволяет избежать ошибки
                settings[name.strip()] = value.strip()
        return settings
    try:
        settings = read_settings(settings_file)
        color=settings.get('theme')
        model=settings.get('model')
        speed=settings.get('speed')
        sg.theme(color)
        if settings.get('debug') == '1':
            debug_mode = True
    except:
        pass

    def send_at_command(port, debug=False):
        # Функция для отправки команды AT на указанный COM порт и получения ответа.
        # Возвращает ответ на команду или None, если ответа нет.
        try:
            ser = serial.Serial(port, timeout=2)
            ser.write(b'AT\r\n')
            response = ser.read(100).decode('utf-8').strip()
            ser.close()
            return response
        except serial.SerialException:
            return None

    modem_port = None

    def check_sms_symbols(message):
        # Проверяет текст на наличие недопустимых символов для текстового режима
        allowed_chars = set(
            'abcdefghijklmnopqrstuvwxyz'
            'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
            '0123456789'
            ' .,!?()-+=:;@')

        for char in message:
            if char not in allowed_chars:
                return False
        return True
    def find_modem():
        global modem_port,debug_mode, speed, model

        # Находим все доступные COM порты
        available_ports = list_ports.comports()

        if not available_ports:
            print("Не удалось найти модем.")
            print('Функции отправки и принятия СМС не будут работать.')
            modem_port = "COM"
        else:
            # Проходим по каждому доступному порту
            for port_info in available_ports:
                port = port_info.device
                device_name = port_info.description  # Получаем имя устройства
                if model in device_name:
                    response = send_at_command(port, "AT")
                    if response:
                        # Сохраняем первый найденный порт и завершаем выполнение
                        modem_port = port
                        break
                else:
                    modem_port = "COM"
            else:
                modem_port = 'COM'

    from gsmmodem.modem import GsmModem

    def send(message, recipient_numbers, pdu):
        global contacts_window, modem_port, speed

        modem = GsmModem(modem_port, speed)
        use_text_mode = check_sms_symbols(message)  # use PDU mode
        if not use_text_mode:
            if not do_continue("Сообщение содержит нестандартные символы, отправить в PDU-режиме?"):
                return
        print("Отправка в PDU-режиме." if not use_text_mode else "Отправка в текстовом режиме.")
        print("\nОтправка сообщений\n[", end="")
        contacts_window.refresh()

        modem.smsTextMode = use_text_mode

        modem.connect("")
        for recipient_number in recipient_numbers:
            modem.sendSms(recipient_number, message)
            print("#", end="")
            contacts_window.refresh()
        print("]")
        contacts_window.refresh()
        modem.close()
        modem = GsmModem(modem_port, speed)
        modem.connect("")
        modem.smsTextMode = True
        modem.close()

    from datetime import timedelta
    warnings.simplefilter(action='ignore', category=FutureWarning)

    def load_contacts(filename):
        try:
            df = pd.read_excel(filename)
            df = df.drop_duplicates()
        except Exception as e:
            print(f"Ошибка при чтении файла контактов: {e}")
            return {}

        required_columns = [df.columns[0], df.columns[1]]  # Номер телефона, Имя
        for column in required_columns:
            if column not in df.columns:
                print(f"Отсутствует ожидаемый столбец: '{column}'")
                return {}

        contacts = {}
        for index, row in df.iterrows():
            phone_number = str(row[df.columns[0]]).replace("а", '%').replace(' ', '').replace('-', '')
            print(f"Phone nu = {phone_number}")
            contacts[phone_number] = phone_number  # row[df.columns[1]]
        return contacts

    def load_sms_log(filename):
        try:
            df = pd.read_excel(filename)
            df = df.drop_duplicates()
            df['Сообщение'] = df['Сообщение'].astype(str)
            #print("Загруженные столбцы:", df.columns)
        except Exception as e:
            print(f"Ошибка при чтении файла SMS логов: {e}")
            return pd.DataFrame()
        return df

    def delete_contact(nums):
        ii = 0
        try:
            # Загружаем файл
            wb = load_workbook("Files/contacts.xlsx")
            ws = wb.active
            # находим все для удаления в обратном порядке
            rows_to_delete = []
            for row in range(ws.max_row, 1, -1):  # начинаем с конца, пропускаем заголовок
                if f"+7{ws.cell(row=row, column=1).value}" in nums:
                    ii += 1
                    rows_to_delete.append(row)

            #удаляем
            for row_idx in rows_to_delete:
                ws.delete_rows(row_idx, 1)

            #Сохраняем
            wb.save("Files/contacts.xlsx")
            return True, f"{ii} Контактов успешно удалено"

        except Exception as e:
            return False, f"Ошибка при удалении контакта: {str(e)}"

    def get_current_datetime():
        now = datetime.now()
        return now.strftime('%d/%m/%Y'), now.strftime('%H:%M:%S')

    def analyze_sms_log(contacts_file, sms_log_file, analysis_file):
        contacts = load_contacts(contacts_file)
        sms_log = load_sms_log(sms_log_file)

        if sms_log.empty or not contacts:
            print("Не удалось загрузить данные для анализа.")
            return

        #Получение даты и времени
        today_date, current_time = get_current_datetime()
        yesterday_date = (datetime.now() - timedelta(days=1)).strftime('%d/%m/%Y')

        #Используем индексаци
        phone_column_index = 0  # Индекс столбца с номерами телефонов
        date_column_index = 3  # Индекс столбца с датой получения SMS

        #Проверка типов
        sms_log.iloc[:, phone_column_index] = sms_log.iloc[:, phone_column_index].astype(str)
        sms_log.iloc[:, date_column_index] = pd.to_datetime(sms_log.iloc[:, date_column_index], format='%d/%m/%Y',
                                                            errors='coerce').dt.strftime('%d/%m/%Y')

        #получаем номера телефонов, которые отправляли SMS за последние сутки
        recent_sms = sms_log[sms_log.iloc[:, date_column_index].isin([today_date, yesterday_date])]
        recent_sms_numbers = recent_sms.iloc[:, phone_column_index].str.replace(' ', '').str.replace('-', '').replace(
            "+7", "").unique()

        #определяем контакты, которые не прислали SMS за последние сутки
        missing_contacts = {number: name for number, name in contacts.items() if number not in recent_sms_numbers}

        #определяем номер анализа
        try:
            with open(analysis_file, 'r', encoding='utf-8') as f:
                lines = f.readlines()
                last_analysis_number = None
                for line in reversed(lines):
                    if line.startswith('Анализ номер '):
                        last_analysis_number = int(line.strip().split('Анализ номер ')[-1].split()[0])
                        break
                if last_analysis_number is None:
                    new_analysis_number = 1
                else:
                    new_analysis_number = last_analysis_number + 1
        except FileNotFoundError:
            new_analysis_number = 1

        # Запись результатов в файл
        analysis_content = f"Анализ номер {new_analysis_number}\n"
        analysis_content += f"Дата: {today_date}, время: {current_time}.\n"
        analysis_content += "Контакты, не приславшие сообщение за последние сутки:\n\n"
        for number, name in missing_contacts.items():
            if len(number) == 11 and not number.startswith('8'):
                analysis_content += f"+{number} -- {name}\n"
            else:
                analysis_content += f"{number} -- {name}\n"

        # Анализ сообщений
        settings = read_settings("Files/settings.txt")
        charge_warning = int(settings.get('charge_warning', 0))
        wb = load_workbook("Files/sms_log.xlsx", data_only=True)
        ws = wb.active

        # Очистка 7-го столбца
        for row in ws.iter_rows(min_row=2, values_only=False):
            row[5].value = ""

        # Добавляем столбец "Отклонения", если его нет
        if ws.max_column < 6:
            ws['G1'] = 'Отклонения'

        for row in ws.iter_rows(min_row=2, values_only=False):
            message = row[2].value
            if isinstance(message, str):  # Проверяем, является ли сообщение строкой
                deviations = row[5].value if row[5].value else ""
                battery_warning = False
                gps_warning = False
                for line in message.splitlines():
                    if "Спутн: 0" in line:
                        gps_warning = True
                    if "Бат:" in line:
                        battery_level = int(line.split("(")[1].split("%")[0])
                        if battery_level < charge_warning:
                            battery_warning = True
                if battery_warning:
                    deviations += "Бат! "
                if gps_warning:
                    deviations += "GPS! "
                row[5].value = deviations

                # Устанавливаем цвет фона для 7-го столбца
                if "Бат! GPS! " in deviations:
                    row[5].fill = PatternFill(start_color='FFFF950E', end_color='FFFF950E', fill_type='solid')
                elif "GPS! " in deviations:
                    row[5].fill = PatternFill(start_color='FFF0F076', end_color='FFF0F076', fill_type='solid')
                elif "Бат! " in deviations:
                    row[5].fill = PatternFill(start_color='FFAFEEEE', end_color='FFAFEEEE', fill_type='solid')

        # Установка белой заливки для пустых ячеек в 7-м столбце
        for row in ws.iter_rows(min_row=2, values_only=False):
            if row[5].value is None or row[5].value == "":
                row[5].fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

        wb.save(sms_log_file)

        with open(analysis_file, 'a', encoding='utf-8') as f:
            f.write("\n\n")
            f.write(analysis_content)

        print(f"Анализ номер {new_analysis_number} успешно добавлен в файл {analysis_file}.")

    def analysis():
        # анализирует
        contacts_file = "Files/contacts.xlsx"
        sms_log_file = "Files/sms_log.xlsx"
        analysis_file = "Files/Analysis.txt"
        analyze_sms_log(contacts_file, sms_log_file, analysis_file)


    def add_contacts(file_path, new_contacts):
        # создаем каталог, если он не существует
        directory = os.path.dirname(file_path)
        if not os.path.exists(directory):
            os.makedirs(directory)

        if os.path.exists(file_path):
            # загружаем существующий файл
            wb = load_workbook(file_path)
            ws = wb.active
        else:
            # создаем новый файл и заполняем заголовки
            wb = Workbook()
            ws = wb.active
            ws.title = "Contacts"
            ws.append(["Phone Number", "Contact Name"])

        # добавляем новые контакты
        for contact in new_contacts:
            ws.append(contact)

        wb.save(file_path)
        print(f"Контакты успешно добавлены в {file_path}")

    def search_contacts(file_path, search_terms):
        wb = load_workbook(file_path)
        ws = wb.active

        if search_terms == "":
            search_terms = ["9", "8", "7", "6", "5", "4", "3", "2", "1", "0"]
        else:
            search_terms = search_terms.split()

        final_strings = []

        for search_term in search_terms:
            if not search_term.startswith("-"):
                for row in ws.iter_rows(min_row=2, values_only=True):
                    phone_number, contact_name = row
                    if phone_number:
                        string = f"+7{phone_number}::{contact_name}".replace(" ", "")

                        if search_term in string and string not in final_strings:
                            final_strings.append(string)
            else:
                argument = search_term.replace("-", "")
                for final_string in final_strings:
                    if argument in final_string:
                        final_strings.remove(final_string)

        contacts_found = []

        for final_string in final_strings:
            num1, name1 = final_string.split("::")
            contacts_found.append({"num": num1, "name": name1})

        if not contacts_found:
            print("Нет контактов, соответствующих критериям поиска.")
            return ["Нет контактов, соответствующих критериям поиска."], []

        final = []
        just_info = []

        for i, contact in enumerate(contacts_found):
            just_info.append({"number": contact["num"], "name": contact["name"]})

        return final, just_info

    import subprocess
    import platform
    def open_files_folder():
        # Открывает папку 'Files' в текущем каталоге в проводнике.

        try:
            # Определяем путь к папке 'Files'
            current_directory = os.getcwd()
            folder_path = os.path.join(current_directory, 'Files')

            # Проверяем, существует ли папка
            if not os.path.isdir(folder_path):
                print(f"Папка не существует: {folder_path}")
                return

            # Открываем папку в проводнике
            if os.name == 'nt':  # Windows
                subprocess.run(['explorer', folder_path], check=True)
            elif os.name == 'posix':  # macOS or Linux
                if sys.platform == 'darwin':  # macOS
                    subprocess.run(['open', folder_path], check=True)
                else:  # Linux
                    subprocess.run(['xdg-open', folder_path], check=True)
            else:
                print(f"Операционная система не поддерживается: {os.name}")

        except subprocess.CalledProcessError as e:
            print('', end='')
        except Exception as e:
            print('', end='')

    from datetime import datetime
    import pandas as pd
    from openpyxl import Workbook
    def send_at_command0(ser, command, response_timeout=1):
        ser.write((command + '\r\n').encode())
        time.sleep(response_timeout)
        response = ser.read_all().decode()
        return response

    # Функция для объединения длинных сообщений
    def combine_long_messages(messages):
        combined_messages = []
        for message in messages:
            combined_messages.append(message)
        return combined_messages

    def num_to_name(num):
        # преобразует номер в имя
        wb = load_workbook("Files/contacts.xlsx")
        ws = wb.active
        # print(f"Искомый номер: {num}")
        for row in ws.iter_rows(min_row=2, values_only=True):
            phone_number, contact_name = row
            if phone_number:
                phone_number = f"+7{phone_number}".replace(" ", "")
                # print(f"Номер контакта: {phone_number}")

                if phone_number == num:
                    return contact_name
        return num

    # Функция для парсинга ответа AT+CMGL и извлечения SMS сообщений
    def parse_sms_response(response):
        messages = []
        lines = response.splitlines()
        i = 0
        '''
        пример сообщения: 
        +CMGL: 10,"REC READ","+79875324724",,"24/11/15,14:35:51+12"
        Hello!
        Или:

        '''
        while i < len(lines):
            if "+CMGL: " in lines[i]:
                parts = lines[i].split(",")
                index = parts[0].split(": ")[1].strip()
                sender_number = parts[2].strip('"')
                date_and_time = lines[i].split(",,")[1].replace('"', '').split(',')
                # print(f"{date_and_time=}")
                date_dates = date_and_time[0].split("/")
                date_date = f"{date_dates[2]}.{date_dates[1]}.{date_dates[0]}"
                # print(f"ДАТА = {date_date}")

                date_time = date_and_time[1].split("+")[0].split("-")[0]
                # print(f"ВРЕМЯ = {date_time}")

                message_lines = []

                # Проверяем, есть ли следующая строка
                if i + 1 < len(lines):
                    # Если следующая строка не начинается с "+CMGL: ", то добавляем ее к сообщению
                    j = i + 1
                    while j < len(lines) and "+CMGL: " not in lines[j]:
                        if "OK" in lines[j]:
                            break
                        message_lines.append(lines[j].strip())
                        j += 1
                    i = j - 1  # Установим индекс на последнюю строку сообщения

                # Декодируем строки сообщения
                decoded_lines = []
                for line in message_lines:
                    try:
                        # Преобразуем сообщение в формат UCS2
                        decoded_line = bytes.fromhex(line).decode('utf-16be')
                        decoded_lines.append(decoded_line)
                    except (ValueError, UnicodeDecodeError):
                        # Если декодирование не удалось, оставляем все строки сообщения в первоначальном виде
                        decoded_lines = message_lines
                        break

                message = '\n'.join(decoded_lines)

                # Преобразуем дату в формат DD/MM/YYYY
                # formatted_date = format_date(date.strip())

                messages.append({
                    "index": index,
                    "sender_number": sender_number,
                    "date": date_date,
                    "time": date_time,
                    "message": message.strip()
                })
            i += 1
        return messages

    def read_sms_and_save(port, contacts_file, output_file):
        global contacts_window, speed
        with serial.Serial(port,speed, timeout=1) as ser:
            # print("Проверяем...")
            response = send_at_command0(ser, 'AT+CMGL="ALL"')

            # Обработка ответа и запись в Excel
            sms_messages = parse_sms_response(response)
            combined_messages = combine_long_messages(sms_messages)

            # Проверяем, существует ли файл с контактами
            if not os.path.exists(contacts_file):
                print(f"Файл {contacts_file} не найден.")
                return

            contacts = load_contacts(contacts_file)

            # Вывод содержимого SMS
            if combined_messages:
                # print()
                # print("Найдены SMS сообщения:", end = '')
                latest_name = "test4d!"
                log = ""
                for sms in combined_messages:
                    # print('')
                    name = num_to_name(sms['sender_number'])
                    log += f"{num_to_name(sms['sender_number'])}: {sms['message']}  {sms['time']}\n"
                    if name != latest_name:
                        print(f">> {num_to_name(sms['sender_number'])}: \n{sms['message']}")
                    else:
                        print(f"{sms['message']}")
                    latest_name = name
                    contacts_window.refresh()
                append_to_excel(combined_messages, contacts, output_file)
                # print("Добавлено, удаляем")
                # Удаление
                # по индексу
                for sms in combined_messages:
                    # print(f"удаляем {sms}")
                    send_at_command0(ser, f"AT+CMGD={sms['index']}")
                return log
            else:
                cy = 1
                if cy == 15:
                    cy = 1
                return ""

    # Функция для загрузки контактов из файла Excel
    def load_contacts(filename):
        try:
            df = pd.read_excel(filename)
        except Exception as e:
            print(f"Ошибка при чтении файла контактов: {e}")
            return {}

        required_columns = ['Номер телефона', 'Имя маячка']
        for column in required_columns:
            if column not in df.columns:
                print(f"Отсутствует ожидаемый столбец: '{column}'")
                return {}

        contacts = {}
        for index, row in df.iterrows():
            # print(row['Номер телефона'])
            # Приведение номеров телефонов к строковому формату без лишних символов
            phone_number = str(row['Номер телефона']).replace(' ', '').replace('-', '')
            contacts[phone_number] = row['Имя маячка']
        return contacts

    from openpyxl.styles import Alignment, PatternFill
    def append_to_excel(sms_messages, contacts, output_file):
        if not sms_messages:  # Если нет новых сообщений, не записываем в таблицу
            return
        try:
            wb = load_workbook("Files/sms_log.xlsx", data_only=True)
            ws = wb.active
        except FileNotFoundError:
            wb = Workbook()
            ws = wb.active
            ws.append(["Номер отправителя", "Имя контакта", "Сообщение", "Дата получения", "Время получения"])

        settings = read_settings("Files/settings.txt")
        sleep_time = int(settings.get('sleep_time', 0))  # Значение sleep_time из файла настроек

        for sms in sms_messages:
            sender_number = sms["sender_number"].replace(' ', '').replace('-',
                                                                          '')  # Убираем только пробелы и дефисы, сохраняем +
            contact_name = num_to_name(sms['sender_number'])
            message = sms["message"] if sms["message"] else "Без текста"
            date_received = sms["date"]
            current_date = datetime.now().strftime('%d/%m/%Y')
            current_time = datetime.now().strftime('%H:%M:%S')

            # Ищем существующую строку с таким же номером и временем
            existing_row = None
            for row in ws.iter_rows(min_row=2, values_only=False):
                if (row[0].value == sender_number and
                        row[3].value == date_received and
                        abs((datetime.strptime(current_time, '%H:%M:%S') - datetime.strptime(row[4].value,
                                                                                             '%H:%M:%S')).total_seconds()) <= sleep_time + 30):
                    existing_row = row
                    break

            if existing_row:
                # Если найдена существующая строка, добавляем к ней новое сообщение
                existing_row[2].value += "\n" + message
                # Увеличиваем высоту строки
                lines = existing_row[2].value.count('\n') + 1
                ws.row_dimensions[existing_row[0].row].height = 13.7
            else:
                # Если не найдена существующая строка, добавляем новую
                ws.append([sender_number, contact_name, message, date_received, current_time])
                # Устанавливаем высоту строки
                lines = message.count('\n') + 1
                ws.row_dimensions[ws.max_row].height = 13.7

        wb.save(output_file)
    import sys

    import serial
    import time
    from openpyxl import load_workbook
    import os
    # from com_port_checker import *

    import shutil
    from datetime import datetime

    def clear_logs():
        log_file = "Files/sms_log.xlsx"
        analysis_file = "Files/Analysis.txt"
        archive_dir = "Files/Archive"
        if not os.path.exists(archive_dir):
            os.makedirs(archive_dir)

        current_datetime = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        archived_log_file = f"{archive_dir}/sms_log_{current_datetime}.xlsx"
        archived_analysis_file = f"{archive_dir}/Analysis_{current_datetime}.txt"

        try:
            shutil.copy2(log_file, archived_log_file)
            print(f"Лог успешно архивирован в {archived_log_file}")

            wb = load_workbook("Files/sms_log.xlsx", data_only=True)
            ws = wb.active
            # Удаляем все строки, кроме первой
            ws.delete_rows(2, ws.max_row)

            # Устанавливаем высоту первой строки по умолчанию
            ws.row_dimensions[1].height = 13.7

            # Удаляем все настройки высоты строк, чтобы они стали по умолчанию
            for row_dim in ws.row_dimensions:
                if row_dim != 1:
                    ws.row_dimensions[row_dim].height = None

            wb.save(log_file)
            print("Лог успешно очищен, кроме строки заголовков")

            # Архивирование файла analysis.txt
            shutil.copy2(analysis_file, archived_analysis_file)
            print(f"Файл analysis.txt успешно архивирован в {archived_analysis_file}")

            # Очистка файла analysis.txt
            with open(analysis_file, 'w') as f:
                f.write('')

            print("Файл analysis.txt успешно очищен")

        except Exception as e:
            print(f"Ошибка при очистке логов: {e}")

    def read_settings(settings_file):
        if not os.path.exists(settings_file):
            print(f"Файл {settings_file} не существует.")
            exit()
        settings = {}
        with open(settings_file, 'r') as file:
            for line_num, line in enumerate(file, start=1):
                line = line.strip()
                if not line or '=' not in line:
                    continue
                try:
                    name, value = line.split('=', 1)
                    settings[name.strip()] = value.strip()
                except ValueError as e:
                    print(f"Ошибка в строке {line_num}: '{line}', ошибка: {e}")
                    continue
        return settings

    def restart_modem():
        global modem_port, speed
        with serial.Serial(modem_port, speed, timeout=1) as ser:
            res = send_at_command0(ser, 'AT+CFUN=1,1')
            return True if "OK" in res else False

    def setup_modem(port):
        global speed
        with serial.Serial(port, speed, timeout=1) as ser:
            send_at_command0(ser, 'AT+CMGF=1')
            send_at_command0(ser, 'AT+CPMS="ME","ME","ME"')
            return "OK"

    def do_continue(text):
        # Затем определяем интерфейс
        layout = [
            [sg.Text(text)],
            [sg.Button('НЕТ'), sg.Button('ДА')]
        ]

        # Создание окна
        window = sg.Window('Продолжить?', layout,  # Запрещает сворачивание
                           keep_on_top=True)

        # Цикл событий
        while True:
            event, values = window.read()

            if event in (sg.WINDOW_CLOSED, "НЕТ"):
                window.close()
                return False

            if event == "ДА":
                window.close()
                return True

    def settings():
        # Читаем текущие настройки
        with open("Files/settings.txt", "r") as f:
            settings = {}
            for line in f:
                if '=' in line:
                    key, value = line.strip().split('=')
                    settings[key.strip()] = value.strip()

        # Получаем список доступных тем
        themes = sg.theme_list()
        current_theme = settings.get('theme', 'DarkAmber')
        current_battery = settings.get('charge_warning', '20')  # По умолчанию 20%
        current_speed = settings.get('speed', '9600')
        current_model = settings.get('model', 'HUAWEI Mobile Connect - 3G PC UI Interface')
        layout = [
            [sg.Text('Тема оформления:')],
            [sg.Combo(themes, default_value=current_theme, key='theme', size=(20, 1))],
            [sg.HSeparator()],
            [sg.Text("Название модема: "), sg.InputText(key='model', default_text=current_model, size=(30, 10), enable_events=True)],
            [sg.Text("Скорость модема: "), sg.InputText(key='speed', default_text=current_speed, size=(30, 10), enable_events=True)],
            [sg.HSeparator()],
            [sg.Text('Уровень заряда для предупреждения:')],
            [sg.Slider(range=(1, 100),
                       default_value=int(current_battery),
                       orientation='h',
                       key='battery',
                       size=(20, 15))],
            [sg.Button("Архивировать даные", key="archive")],
            [sg.HSeparator()],
            [sg.Button('Сохранить'), sg.Button('Отмена')]
        ]

        window = sg.Window('Настройки', layout,
                           disable_minimize=True,  # Запрещает сворачивание
                           keep_on_top=True)

        while True:
            event, values = window.read()

            if event in (sg.WIN_CLOSED, 'Отмена'):
                break

            if event == "archive":
                clear_logs()

            if event == 'Сохранить':
                # Сохраняем настройки
                settings['theme'] = values['theme']
                settings['charge_warning'] = str(int(values['battery']))
                settings['speed'] = values['speed']
                settings['model'] = values['model']
                with open("Files/settings.txt", "w") as f:
                    for key, value in settings.items():
                        f.write(f"{key} = {value}\n")

                # Обновляем тему
                sg.theme(values['theme'])
                with open("Files/color.txt", "w") as f:
                    f.write(values['theme'])

                break

        window.close()

    def menu_analysing():
        if do_continue("Анализировать данные? 🤨"):
            analysis()
            err_msg("Успешно 👌")

    def kill_connect_manager():
        try:
            # Ищем процесс Connect Manager
            for proc in psutil.process_iter(['name']):
                if proc.info['name'] and 'Connect Manager.exe' in proc.info['name']:
                    # print(f"Найден процесс Connect Manager (PID: {proc.pid})")
                    # Принудительно завершаем процесс
                    proc.kill()
                    # print("Процесс успешно завершен")
                    return True

            # print("Процесс Connect Manager.exe не найден")
            return False

        except Exception as e:
            print(f"Произошла ошибка: {e}")
            return False

    def timer(seconds: int):
        # Создаем окно с таймером
        layout = [
            [sg.Text('Сколько осталось ждать:', font='Helvetica 12')],
            [sg.Text('', size=(10, 1), font='Helvetica 20 bold', key='timer')],
            # [sg.Button('Отмена', font='Helvetica 10')]
        ]

        window = sg.Window('Таймер', layout, finalize=True, no_titlebar=True,
                           disable_minimize=True,  # Запрещает сворачивание
                           keep_on_top=True,  # Держит окно поверх других
                           grab_anywhere=True
                           )

        # Запускаем таймер
        start_time = time.time()
        remaining = seconds

        while remaining > 0:
            event, values = window.read(timeout=100)  # Обновляем каждые 100мс

            if event in (sg.WIN_CLOSED, 'Отмена'):
                window.close()
                return False

            # Обновляем оставшееся время
            current_time = time.time()
            elapsed = int(current_time - start_time)
            remaining = seconds - elapsed

            # Обновляем текст таймера
            window['timer'].update(f'{remaining} сек')

        window.close()
        return True

    def err_msg(text):
        global contacts_window
        global can_modem
        # Затем определяем интерфейс
        layout = [
            [sg.Text(text), sg.Button('Смириться', font='Helvetica 12 bold')]
        ]

        # Создание окна
        window = sg.Window('Уведомление', layout, no_titlebar=True,  # Держит окно поверх других
                           grab_anywhere=True)

        # Цикл событий
        while True:
            event, values = window.read()

            if event in (sg.WINDOW_CLOSED, 'Смириться'):
                break
        window.close()

    kill_connect_manager()
    def reload_data():
        global contacts_data
        # Загружаем существующие контакты
        existing = search_contacts("Files/contacts.xlsx", values["args"])[1]

        #print(f"Контакты = {existing}")

        # Преобразуем существующие контакты в формат для таблицы
        contacts_data = []
        if existing:
            for el in existing:
                contacts_data.append([el["name"], el["number"]])
        contacts_window["table"].update(values=contacts_data)

    selected_numbers = []
    # Загружаем существующие контакты
    existing = search_contacts("Files/contacts.xlsx", "")[1]

    #print(f"Контакты = {existing}")

    # Создаем заголовки для таблицы
    headings = ['Имя', 'Телефон']

    # Преобразуем существующие контакты в формат для таблицы
    contacts_data = []
    if existing:
        for el in existing:
            contacts_data.append([el["name"], el["number"]])
    total_console = ""


    layout = [
        [sg.Text('Имя:', font='Helvetica 12 bold'), sg.InputText(key='name',size=(38,10), font='Helvetica 12 bold'), sg.Button("Получить сообщения", font='Helvetica 12 bold', key="get"), sg.Button("Обновить", font='Helvetica 12 bold',key="update"), sg.Button("⟳", font='Helvetica 12 bold'), sg.Button("ⓘ", font='Helvetica 12 bold')],
        [sg.Text('Телефон:', font='Helvetica 12 bold'), sg.InputText(key='phone', font='Helvetica 12 bold',size=(34,10)), sg.Button('Анализировать данные', font='Helvetica 12 bold'), sg.Button('Настройки', font='Helvetica 12 bold')],
        [sg.Button('Добавить контакт', font='Helvetica 12 bold'), sg.Button('Очистить', font='Helvetica 12 bold'), sg.Button('Удалить', font='Helvetica 12 bold')],
        [sg.Text('Список контактов:', font='Helvetica 12 bold'), sg.Text('Аргументы для поиска: ', font='Helvetica 12 bold'), sg.InputText(key='args',size=(27,10), font='Helvetica 12 bold'), sg.Button("Найти контакты", font='Helvetica 12 bold', key="find", bind_return_key=True)],
        [sg.Table(values=contacts_data,
                 headings=headings,
                 max_col_width=55,
                 col_widths=[5, 11],
                 alternating_row_color="",
                 auto_size_columns=False,
                 justification='left',
                 num_rows=10,
                 key='table',
                 enable_events=True,
                 font = 'Helvetica 12 bold',
                 size=(60, 20),
                 select_mode=sg.TABLE_SELECT_MODE_EXTENDED),sg.Multiline(size=(60, 11), key='menu_console', autoscroll=True, reroute_stdout=True,
                 reroute_stderr=False, font='Helvetica 12 bold', write_only=True, disabled=True,border_width=3)],
        [sg.Button('Выход', font='Helvetica 12 bold'), sg.Text('Сообщение: ', font='Helvetica 12 bold'), sg.InputText(key='msg', font='Helvetica 12 bold', size=(52,10)), sg.Button('Отправить!', font='Helvetica 12 bold')]
    ]

    # Создание окна
    contacts_window = sg.Window('Центр управления сообщениями', layout, icon=r"C:\Users\vlkardakov\Documents\1\Bots\SmsToolsN\social.ico", finalize=True)
    contacts_window.refresh()
    if True:
        find_modem()
        if modem_port != "COM":
            setup_modem(modem_port)
            can_modem = True
    # Цикл событий
    while True:
        event, values = contacts_window.read()

        #print("Окно прочитано ;D")

        #print(event)
        #print(values)

        if event == 'table':  # когда кликаем по таблице

            selected_rows = values['table']
            selected_numbers = []
            ii=0
            for row_index in selected_rows:
                selected_contact = contacts_data[row_index]
                ii+=1
                selected_numbers.append(selected_contact[1])
                contacts_window['name'].update(selected_contact[0])
                contacts_window['phone'].update(selected_contact[1])
            print(f"Выбрано {ii} контактов")


        if event in (sg.WINDOW_CLOSED, 'Выход'):
            break

        if event == 'Настройки':
            settings()
            contacts_window.close()
            menu_contacts()

        if event == '⟳':
            try:
                if do_continue("Перезагрузить модем (50 секунд)?"):
                    res = restart_modem()
                    kill_connect_manager()
                    timer(50)
                    kill_connect_manager()
                    time.sleep(2)
                    setup_modem(modem_port)
            except:
                err_msg("Не удалось перезагрузить модем")


        if event == "ⓘ":
            try:
                open_files_folder()
            except:
                pass
        if event == "Анализировать данные":
            try:
                reload_data()
                menu_analysing()
            except:
                err_msg("Не удалось анализировать данные :/")

        if event == "choose_all":
            ids_to_choose = []
            for i in range(len(contacts_data)):
                ids_to_choose.append(i)
            contacts_window["table"].update(ids_to_choose)
            pass

        if event == 'Добавить контакт':
            try:
                if values['name'] and values['phone']:
                    add_contacts("Files/contacts.xlsx", [[values["phone"].replace("+7", ""),values["name"]]])
                    new_contact = [values['name'], values['phone']]
                    contacts_data.append(new_contact)
                    contacts_window['table'].update(values=contacts_data)

                    reload_data()


                    print(f"Добавлен контакт: {new_contact}")
                    # Очищаем поля ввода
                    contacts_window['name'].update('')
                    contacts_window['phone'].update('')
            except:
                pass

        if event == 'update':
            try:
                kill_connect_manager()
                setup_modem(modem_port)
                reload_data()
                kill_connect_manager()
            except:
                err_msg("Не удалось подключиться к модему.")

        if event == 'find':
            try:
                reload_data()
            except:
                pass
        if event == "Удалить":
            try:
                if selected_numbers:
                    selected_numbers_count = len(selected_numbers)
                    delete_contacts_message = ""
                    if 5 <= selected_numbers_count%10 or (10 < selected_numbers_count < 21):
                        delete_contacts_message=f"Удалить {selected_numbers_count} контактов?"
                    elif selected_numbers_count%10==1:
                        delete_contacts_message=f"Удалить {selected_numbers_count} контакт?"
                    elif 1 < selected_numbers_count%10 < 5:
                        delete_contacts_message = f"Удалить {selected_numbers_count} контакта?"

                    if do_continue(delete_contacts_message):
                        delete_contact(selected_numbers)
                        reload_data()
                        err_msg("Успешно.")
                else:
                    err_msg("Сначала выберите контакты!")
            except:
                pass
        if event == 'Отправить!' and values["msg"] and do_continue("Отправить сообщение?"):
            try:
                if can_modem:
                    if selected_numbers:
                        send(values["msg"], selected_numbers, False)
                        print(f"Сообщения отправлены! :D")
                        time.sleep(0.1)
                    else:
                        err_msg("Сначала выберите контакты!")
                else: err_msg("Модем не подключен.")
            except:
                err_msg("Не удалось отправить сообщения :/")
        if event == 'Очистить':
            contacts_window['name'].update('')
            contacts_window['phone'].update('')
        if event == "get":
            if can_modem:
                read_sms_and_save(modem_port, contacts_file, output_file)
            else:
                err_msg("Модем не подключен.")


    contacts_window.close()
if __name__ == "__main__":
    print(len(sg.theme_list()))
    menu_contacts()
